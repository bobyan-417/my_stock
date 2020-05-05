from openpyxl import load_workbook
from _sqlite3 import Row
import string
# -*- coding: utf-8 -*-

class ExcelHandler(object):
    def __init__(self, filename):
        self.filename = filename

    def tostr(self, s): # Clean out non-unicode chars for csv.writer - SLOW
        try:
            return str(s)
        except:
            return s
    
    def getAllValues(self, sheetname):
        wb = load_workbook(filename = self.filename, use_iterators=True)
        worksheet = wb.get_sheet_by_name(sheetname)
        lines = []
        for row in worksheet.iter_rows():
            values = []
            for cell in row:
                values.append(self.tostr(cell.value))
            lines.append(values)
        return lines
            



if __name__ == '__main__': 
    
    excel = ExcelHandler('C:/bobyan/project/Stock/Shanghai.xlsx')
    lines = list(excel.getAllValues("sz"))
    for line in lines:        
        print(line[0] + ":" + line[1])
        
        



